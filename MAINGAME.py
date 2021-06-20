
import tkinter as tk
from tkinter import PhotoImage, font as tkfont
from tkinter.constants import END, LEFT, RIGHT
from typing import Text
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
import AMGAME
import PMGAME
from PMGAME import *



#this part is where the frames are staged and change
class SampleApp(tk.Tk):

    def __init__(self,*args, **kwargs):
        tk.Tk.__init__(self,*args, **kwargs)

        self.title_font = tkfont.Font(family='Helvetica', size=18, weight="bold", slant="italic")

        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        #container.grid_rowconfigure(0, weight=1)
        #container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (startwindow,PMGAME.PMGame,pmendwindow,pmwinwindow,AMGAME.AMGame,amendwindow,amwinwindow,loginpage):
            page_name = F.__name__
            frame = F(master=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("loginpage")


    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()


class loginpage(tk.Frame):

    def __init__(self,master,controller):
        tk.Frame.__init__(self,master)
        self.controller = controller
        self.strvar = tk.StringVar()
        self.label1 = tk.Label(self,textvariable=self.strvar)
        self.Label4 = tk.Label(self, text='Enter your name')
        self.Label4.pack(side=tk.TOP)
        self.label1.pack(side=tk.BOTTOM)
        self.player_name = tk.Entry(self,bd=5)
        self.player_name.pack(side=tk.TOP)
        self.submitbutton = tk.Button(self,text='Submit',width=10,command=self.checkEntry)
        self.submitbutton.pack()
    def checkEntry(self):
        self.name = self.player_name.get()
        if self.name  != "":
            self.controller.show_frame("startwindow")
        else:
            self.strvar.set("Please enter in a name")
#this is the starting window
class startwindow(tk.Frame):

    def __init__(self,master,controller):
        tk.Frame.__init__(self,master)

        self.controller = controller 
        #self.losetext =tk.StringVar()

        #Define the function for the timer
        #canvas = tk.Canvas(self, width = 800, height = 400)
        #canvas.pack()

       # tk.img = PhotoImage(file="image.png")
       # canvas.create_image(20,20,anchor=tk.NW, image=tk.img)

    
        self.label1 = tk.Label(self, text="This is training for both AM and PM",font='Helvetica 12')
        self.label2 = tk.Label(self, text="HOW TO PLAY: First Pick the correct Game and get the highest score you can", font='Helvetica 10')
        self.label3 = tk.Label(self, text="WARING: if you get to many wrong the game will reset  ", font='Helvetica 10')
        #self.label4 = tk.Label(self, textvariable=self.losetext)
        self.Button1 = tk.Button(self,text="PM",command=lambda: self.controller.show_frame("PMGame"),width=10,height=2,bg="gold")
        self.Button2 = tk.Button(self,text="AM",command=lambda: self.controller.show_frame("AMGame"),width=10,height=2,fg='white',bg="blue4")

  
        #self.quitButton = tk.Button(self, text="  QUIT  ",width= 10, height= 2 ,command=self.frame.quit,bg="red")
        #self.quitButton.pack(side=tk.BOTTOM)
        self.Button1.pack(side=tk.BOTTOM)
        self.Button2.pack(side=tk.BOTTOM)
        self.label1.pack(side=tk.TOP,padx=15, pady=2)
        self.label2.pack(side=tk.TOP,padx=15,pady=2)
        self.label3.pack(side=tk.TOP,padx=15,pady=2)

        #self.label4.pack(side=tk.BOTTOM)

#this is the end window for the AM game 
class amendwindow(tk.Frame):

    def __init__(self,master,controller):
        tk.Frame.__init__(self,master)
        #into = tracgame(master,controller)
        self.controller = controller
        #self.score = into.yourscore
        #self.label2 = tk.Label(self, text="Your score was " + str(self.score))
        self.label1 = tk.Label(self, text="OOPS!!!! you got to many mistake but you can retry")
        self.Button1 = tk.Button(self,text="RETRY",command=lambda: self.controller.show_frame("AMGame"),width=10,height=2,bg="green")
        self.label1.pack()
        #self.label2.pack()
        self.Button1.pack()

#this is the end window for the PM game
class pmendwindow(tk.Frame):

    def __init__(self,master,controller):
        tk.Frame.__init__(self,master)
        #into = tracgame(master,controller)
        self.controller = controller
        #self.score = into.yourscore
        #self.label2 = tk.Label(self, text="Your score was " + str(self.score))
        self.label1 = tk.Label(self, text="OOPS!!!! you got to many mistake but you can retry")
        self.Button1 = tk.Button(self,text="RETRY",command=lambda: self.controller.show_frame("PMGame"),width=10,height=2,bg="green")
        self.label1.pack()
        #self.label2.pack()
        self.Button1.pack()

#this is the window for getting all area code correct 
class pmwinwindow(tk.Frame):

    def __init__(self,master,controller):
        tk.Frame.__init__(self,master)
        self.info = PMGAME.PMGame(master,controller)
        self.info2 = loginpage(master,controller)
        self.controller = controller
        #into = tracgame(master,controller)
        #self.score = into.yourscore
        #self.label2 = tk.Label(self, text="Your score was " + str(self.score))
        self.label1 = tk.Label(self, text="good job you know all the area codes !!!!!!!!!!!!. You can Continue or Quit")
        self.Button1 = tk.Button(self,text="Continue",command=lambda: self.controller.show_frame("PMGame"),width=10,height=2,bg="green")
        self.Button2 = tk.Button(self,text="No I wish to Quit",command=self.upload)
        self.label1.pack()
        #self.label2.pack()
        self.Button1.pack()
        self.Button2.pack()
    def upload(self):
        date = self.info.now
        score = self.info.score
        wb = load_workbook('MASTEREXCEL.xlsx')
        ws = wb['USERS']
        ws.append({'B': score,'C': date})
        wb.save("MASTEREXCEL.xlsx")
        
        
        



#this is the window for getting all zip codE correct///
class amwinwindow(tk.Frame):

    def __init__(self,master,controller):
        tk.Frame.__init__(self,master)
        #into = tracgame(master,controller)
        self.controller = controller
        #self.score = into.yourscore
        #self.label2 = tk.Label(self, text="Your score was " + str(self.score))
        self.label1 = tk.Label(self, text="good job you know all the area codes !!!!!!!!!!!!. You can Continue or Quit")
        self.Button1 = tk.Button(self,text="Continue",command=lambda: self.controller.show_frame("AMGame"),width=10,height=2,bg="green")
        self.label1.pack()
        #self.label2.pack()
        self.Button1.pack()



# this runs the whole program 
def main():
    app = SampleApp()
    app.geometry("800x400")
    app.mainloop()
if __name__ == '__main__':
    main()