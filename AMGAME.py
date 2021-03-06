import time
import random
import tkinter as tk
from tkinter import PhotoImage, font as tkfont
from tkinter.constants import END, LEFT, RIGHT
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

class AMGame(tk.Frame):

    def __init__(self, master,controller):
        tk.Frame.__init__(self,master)
        # Global Variables
        # When calling these variables use .self first 
        wb = load_workbook('MASTEREXCEL.xlsx')
        wsAM = wb["AMZIPCODES"]
        self.controller = controller
        self.lives = 3
        self.score = 0
        self.yourscore = 0
        self.highscore = 0
        self.index = 0 
        #this work for getting the values from the excel sheet and removes the none values 
        self.GreenBelt = []
        for col in wsAM['A']:
             self.GreenBelt.append(col.value)
        self.GreenBelt = [i for i in self.GreenBelt if i != None]

        self.BlueBelt = []
        for col in wsAM['B']:
             self.BlueBelt.append(col.value)
        self.BlueBelt = [i for i in self.BlueBelt if i != None]

        self.RedBelt = []
        for col in wsAM['C']:
             self.RedBelt.append(col.value)
        self.RedBelt = [i for i in self.RedBelt if i != None]

        self.YellowBelt = []
        for col in wsAM['D']:
             self.YellowBelt.append(col.value)
        self.YellowBelt = [i for i in self.YellowBelt if i != None]

        self.BlowBy = []
        for col in wsAM['E']:
             self.BlowBy.append(col.value)
        self.BlowBy = [i for i in self.BlowBy if i != None]

        self.Other = []
        for col in wsAM['F']:
             self.Other.append(col.value)
        self.Other = [i for i in self.Other if i != None]

        self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
        self.choice = random.choice(self.fullList)
        
        # This is the GUI 
        # ---------------------------------------
        # This parts makes the frame and for updateing labels 

        self.highscoretext = tk.StringVar()

        self.text = tk.StringVar()

        self.Lives = tk.StringVar()
        self.Lives.set(" You have " + (str(self.lives)) + " Lives left")

        self.pickchoice = tk.StringVar()
        self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))

        self.numberscore = tk.StringVar()
        self.numberscore.set("Your Score: " + str(self.score))
        # Main parts of the Gui that you see
        self.Score = tk.Label(self, textvariable=self.numberscore,font='Helvetica 10 bold')
        self.Score.pack(side=tk.TOP, pady=1, padx=1)

        self.label3 = tk.Label(self, textvariable=self.Lives, font='Helvetica 10 bold')
        self.label3.pack(side=tk.TOP, pady=1, padx=1)

        self.label2 = tk.Label(self, textvariable=self.text)
        self.label2.pack(side=tk.BOTTOM)

        self.label1 = tk.Label(self, textvariable=self.pickchoice, font='Helvetica 15 bold')
        self.label1.pack(side=tk.TOP)

        self.greenbeltbutton = tk.Button(self, text="Green Belt",width= 10, height= 2, bg='gold' ,command=self.GREENBELTCHECK)
        self.greenbeltbutton.pack(side=LEFT, pady=2, padx=15)

        self.bluebeltbutton = tk.Button(self, text="Blue Belt",width= 10, height= 2, fg='white' , bg='blue4' ,command=self.BLUEBELTCHECK)
        self.bluebeltbutton.pack(side=LEFT, pady=2, padx=15)

        self.redbeltbutton = tk.Button(self, text="Red Belt",width= 10, height= 2, bg='gold',command=self.REDBELTCHECK)
        self.redbeltbutton.pack(side=LEFT, pady=2, padx=15)

        self.blowbybutton = tk.Button(self, text="Blow By",width= 10, height= 2, fg='white' , bg='blue4' ,command=self.BLOWBYCHECK)
        self.blowbybutton.pack(side=RIGHT, pady=2, padx=15)

        self.otherbutton = tk.Button(self, text="Other",width= 10, height= 2, bg='gold'  ,command=self.OTHERCHECK)
        self.otherbutton.pack(side=RIGHT, pady=2, padx=15)

        self.yellowbeltbutton = tk.Button(self, text="Yellow Belt",width= 10, height= 2, fg='white' , bg='blue4' ,command=self.YELLOWBELTCHECK)
        self.yellowbeltbutton.pack(side=RIGHT, pady=2, padx=15)

        self.menubutton = tk.Button(self,text="Return To Menu",command=self.MENURETURN)
        self.menubutton.pack()

        # end of GUI
        # -----------------------------------------
    # These are fuctions for checking the answers    
    # These functions will choose a new area code only when you pick the right choice for now  

    def GREENBELTCHECK (self,):
        answer = self.choice
        if answer in self.GreenBelt:
            self.score =  self.score + 1
            self.yourscore = self.score 
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                self.controller.show_frame("amwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))         
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.currentanswer = answer
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                    self.controller.show_frame("amendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
            
    def BLUEBELTCHECK (self,):
        answer = self.choice
        if answer in self.BlueBelt:
            self.score =  self.score + 1
            self.yourscore = self.score 
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")          
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area zip belong: " + str(self.choice))
                self.controller.show_frame("amwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area zip belong: " + str(self.choice))          
        else:
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                    self.controller.show_frame("amendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
          
    def REDBELTCHECK (self):
        answer = self.choice
        if answer in self.RedBelt:
            self.score =  self.score + 1
            self.yourscore = self.score 
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.choice = random.choice(self.fullList)
                self.controller.show_frame("amwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                    self.controller.show_frame("amendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
    
    def YELLOWBELTCHECK (self):
        answer = self.choice
        if answer in self.YellowBelt:
            self.score =  self.score + 1
            self.yourscore = self.score 
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.choice = random.choice(self.fullList)
                self.controller.show_frame("amwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                    self.controller.show_frame("amendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                   
    def BLOWBYCHECK (self):
        answer = self.choice
        if answer in self.BlowBy:
            self.score =  self.score + 1
            self.yourscore = self.score 
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.choice = random.choice(self.fullList)
                self.controller.show_frame("amwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                    self.controller.show_frame("amendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
            
    def OTHERCHECK (self):
        answer = self.choice
        if answer in self.Other:
            self.score =  self.score + 1
            self.yourscore = self.score 
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.choice = random.choice(self.fullList)
                self.controller.show_frame("amwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:   
                self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this zip code belong: " + str(self.choice))
                    self.controller.show_frame("amendwindow")  
            else:  
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
    
    def MENURETURN(self):
        self.lives = 3
        self.score = 0 
        self.fullList = self.GreenBelt + self.BlueBelt + self.RedBelt + self.YellowBelt + self.BlowBy + self.Other
        self.choice = random.choice(self.fullList)
        self.numberscore.set("Your Score: " + str(self.score))
        self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
        self.text.set("")
        self.controller.show_frame("startwindow")

    '''
    just dont worry about this but dont remove its something and nothing at the same time 
    def sqltalk(self):
            self.dbhighscore = sqlengine.mycursor.execute("SELECT highscore FROM REALTEST")
            if int(self.dbhighscore) < self.highscore:
                sqlengine.mycursor.execute("INSERT INTO REALTEST (Fname, HIGHSCORE) VALUES (%s,%s)",('lance',self.highscore))
                sqlengine.db.commit()
    '''