import time
import MAINGAME
import random 
import tkinter as tk
from tkinter import PhotoImage, font as tkfont
from tkinter.constants import END, LEFT, RIGHT
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

class PMGame(tk.Frame):

    def __init__(self, master,controller):
        tk.Frame.__init__(self,master)
        # Global Variables
        # When calling these variables use .self first
        wb = load_workbook('MASTEREXCEL.xlsx')
        wsPM= wb["PMCODES"]
        wsUSER = wb["USERS"]
        self.now = time.strftime("%x")
        self.controller = controller
        self.lives = 3
        self.score = 0
        self.highscore = 0
        self.index = 0
        #this work for getting the values from the excel sheet and removes the none values 

        self.excelscores = []
        for col in wsUSER['B']:
            self.excelscores.append(col.value)
        self.excelscores = [i for i in self.excelscores if i != 'SCORE']
        self.excelscores = [i for i in self.excelscores if i != 'None']
        if self.highscore > 0:
            self.highscore = max(self.excelscores)

        self.SoCal = []
        for col in wsPM['A']:
            self.SoCal.append(col.value)
        self.SoCal = [i for i in self.SoCal if i != None]

        self.NoCal = []
        for col in wsPM['B']:
            self.NoCal.append(col.value)
        self.NoCal = [i for i in self.NoCal if i != None]

        self.Desert = []
        for col in wsPM['C']:
            self.Desert.append(col.value)
        self.Desert = [i for i in self.Desert if i != None]
        
        self.Vancouver = []
        for col in wsPM['D']:
            self.Vancouver.append(col.value)
        self.Vancouver = [i for i in self.Vancouver if i != None]

        self.Seattle = []
        for col in wsPM['E']:
            self.Seattle.append(col.value)
        self.Seattle = [i for i in self.Seattle if i != None]

        self.Other = []
        for col in wsPM['F']:
            self.Other.append(col.value)
            self.Other = [i for i in self.Other if i != None]

        self.fullList = list(self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other)
        self.choice = random.choice(self.fullList)
        
        # This is the GUI 
        # ---------------------------------------
        # This parts makes the frame and for updateing labels 

        self.highscoretext = tk.StringVar()
        self.highscoretext.set("The current high score is: " + str(self.highscore))

        self.text = tk.StringVar()

        self.Lives = tk.StringVar()
        self.Lives.set(" You have " + (str(self.lives)) + " Lives left")

        self.pickchoice = tk.StringVar()
        self.pickchoice.set(" Where does this area code belong: " + str(self.choice))

        self.numberscore = tk.StringVar()
        self.numberscore.set("Your Score: " + str(self.score))
        # Main parts of the Gui that you see
        self.label7 = tk.Label(self, textvariable=self.highscoretext,font='Helvetica 10 bold')
        self.label7.pack(side=tk.TOP, pady=1, padx=1)

        self.Score = tk.Label(self, textvariable=self.numberscore,font='Helvetica 10 bold')
        self.Score.pack(side=tk.TOP, pady=1, padx=1)

        self.label3 = tk.Label(self, textvariable=self.Lives, font='Helvetica 10 bold')
        self.label3.pack(side=tk.TOP, pady=1, padx=1)

        self.label2 = tk.Label(self, textvariable=self.text)
        self.label2.pack(side=tk.BOTTOM)

        self.label1 = tk.Label(self, textvariable=self.pickchoice, font='Helvetica 15 bold')
        self.label1.pack(side=tk.TOP)

        self.socalbutton = tk.Button(self, text="So Cal",width= 10, height= 2, bg='gold' ,command=self.SOCALCHECK)
        self.socalbutton.pack(side=LEFT, pady=2, padx=15)

        self.nocalbutton = tk.Button(self, text="No Cal",width= 10, height= 2, fg='white' , bg='blue4' ,command=self.NOCALCHECK)
        self.nocalbutton.pack(side=LEFT, pady=2, padx=15)

        self.desertbutton = tk.Button(self, text="Desert",width= 10, height= 2, bg='gold',command=self.DESERTCHECK)
        self.desertbutton.pack(side=LEFT, pady=2, padx=15)

        self.seattlebutton = tk.Button(self, text="Seattle",width= 10, height= 2, fg='white' , bg='blue4' ,command=self.SEATTLECHECK)
        self.seattlebutton.pack(side=RIGHT, pady=2, padx=15)

        self.vanbutton = tk.Button(self, text="Vancouver",width= 10, height= 2, bg='gold' ,command=self.VANCOUVERCHECK)
        self.vanbutton.pack(side=RIGHT, pady=2, padx=15)

        self.otherbutton = tk.Button(self, text="Other",width= 10, height= 2, fg='white' , bg='blue4' ,command=self.OTHERCHECK)
        self.otherbutton.pack(side=RIGHT, pady=2, padx=15)
        
        self.menubutton = tk.Button(self,text="Return To Menu",command=self.MENURETURN)
        self.menubutton.pack()

        # end of GUI
        # -----------------------------------------
    # These are fuctions for checking the answers    
    # These functions will choose a new area code only when you pick the right choice for now  

    def NOCALCHECK (self,):
        answer = self.choice
        if answer in self.NoCal:
            self.score =  self.score + 1
            if int(self.score) >= int(self.highscore):
                self.highscoretext.set("The current high score is: " + str(self.score))
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                self.winscore = self.score
                self.controller.show_frame("pmwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))      
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                    self.controller.show_frame("pmendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
            
    def SOCALCHECK (self,):
        answer = self.choice
        if answer in self.SoCal:
            self.score =  self.score + 1
            if int(self.score) >= int(self.highscore):
                self.highscoretext.set("The current high score is: " + str(self.score))
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")          
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                self.winscore = self.score
                self.controller.show_frame("pmwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))          
        else:
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                    self.controller.show_frame("pmendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
        
    def DESERTCHECK (self):
        answer = self.choice
        if answer in self.Desert:
            self.score =  self.score + 1
            if int(self.score) >= int(self.highscore):
                self.highscoretext.set("The current high score is: " + str(self.score))
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                self.winscore = self.score
                self.controller.show_frame("pmwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                    self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                    self.controller.show_frame("pmendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
    
    def VANCOUVERCHECK (self):
        answer = self.choice
        if answer in self.Vancouver:
            self.score =  self.score + 1
            if int(self.score) >= int(self.highscore):
                self.highscoretext.set("The current high score is: " + str(self.score))
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                self.winscore = self.score
                self.controller.show_frame("pmwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                    self.controller.show_frame("pmendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                                  
    def SEATTLECHECK (self):
        answer = self.choice
        if answer in self.Seattle:
            self.score =  self.score + 1
            if int(self.score) >= int(self.highscore):
                self.highscoretext.set("The current high score is: " + str(self.score))
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                self.winscore = self.score
                self.controller.show_frame("pmwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                    self.controller.show_frame("pmendwindow")
            else:
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
            
    def OTHERCHECK (self):
        answer = self.choice
        if answer in self.Other:
            self.score =  self.score + 1
            if int(self.score) >= int(self.highscore):
                self.highscoretext.set("The current high score is: " + str(self.score))
            self.winscore = self.score
            self.numberscore.set("Your Score: " + str(self.score))
            self.text.set("Your Answer is: correct")
            self.fullList.remove(answer)
            if self.fullList == []:
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                self.yourscore = self.score
                self.controller.show_frame("pmwinwindow")
            else:
                self.choice = random.choice(self.fullList)
                self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
        else: 
            self.score =  self.score - 1
            self.lives = self.lives - 1
            if self.score < 0:
                self.score = 0
            if self.lives <= 0 or self.fullList == []:   
                self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
                self.lives = 3
                self.score = 0
                self.numberscore.set("Your Score: " + str(self.score))
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
                if self.lives == 3:
                    self.choice = random.choice(self.fullList)
                    self.text.set("")
                    self.pickchoice.set(" Where does this area code belong: " + str(self.choice))
                    self.controller.show_frame("pmendwindow")  
            else:  
                self.numberscore.set("Your Score: " + str(self.score))
                self.text.set("Your Answer is: not correct")
                self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
    
    def MENURETURN(self):
        self.lives = 3
        self.score = 0 
        self.fullList = self.SoCal + self.NoCal + self.Desert + self.Vancouver + self.Seattle + self.Other
        self.choice = random.choice(self.fullList)
        self.numberscore.set("Your Score: " + str(self.score))
        self.Lives.set(" You have " + (str(self.lives)) + " Lives left")
        self.text.set("")
        self.controller.show_frame("startwindow")
    
    

    ''' dont pay attention to this right here it means nothing but something but not right now
        def sqltalk(self):
        self.dbhighscore = sqlengine.mycursor.execute("SELECT highscore FROM REALTEST")
        if int(self.dbhighscore) < self.highscore:
            sqlengine.mycursor.execute("INSERT INTO REALTEST (Fname, HIGHSCORE) VALUES (%s,%s)",('lance',self.highscore))
            sqlengine.db.commit()
    '''