import time 
import random
from openpyxl.worksheet import worksheet
import pandas as pd 
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
df = pd.read_excel('MASTEREXCEL.xlsx',sheet_name='USERS')
wb = load_workbook('MASTEREXCEL.xlsx')
ws = wb['USERS']    

now = time.strftime("%x")

number = random.randint



x = input('enter your name: ')
y = input('enter your score: ')
test = ws['A']
name = x 
score = y
b = 0


df_name = df[df["NAMES"] == name]

for row in df_name.iterrows():
    b += 1
    ws.insert_rows(idx=1,amount=b)
    ws.append({'B': score,'C': now})
    wb.save("MASTEREXCEL.xlsx")
    
   
# this get all the values from the name col and gets rid of nones

#------------------------------------------
'''
for col in ws['A']:
    names.append(col.value)
    names = [i for i in names if i != None]
 #------------------------------------------
for col in ws['B']:
    score.append(col.value)
    score = [i for i in names if i != None]


if name not in names or newnames :
    newnames.append(name)


for row in ws.iter_rows("A"):
    for any_cell in row:
        if any_cell.value == "bob":
            print(1)






if name in names:
    for row in ws.iter_rows():
        if int(score1) > 30:
            ws.cell(row='2',col='B').value = 'score1'
    
    wb.save('USER_BOOK.xlsx')


       
for i in ws.iter_cols(min_row=1,min_col=1, max_row=6,max_col=3):
        for cell in i:
            print(cell.value, end=" ")
        print()
'''
"""

for i in range(1,ws.max_row):
    if ws.cell(row=row, column=0).value == 'bob':
        for j in range(i, ws.max_column):
            print (ws.cell(row=i, column=j).value)
"""